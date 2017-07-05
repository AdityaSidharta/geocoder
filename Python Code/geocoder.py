import easygui
import sys
import os
import openpyxl
import xlsxwriter
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
import time
from geopy import geocoders

'''
not updated
'''
apispare = ["AIzaSyB2Y_sIIS0JUseEaN3_GBbqIlqpARMbCRc","AIzaSyD6QHap058UQDMA_L9S7OrUYTUUtHwSpFI","AIzaSyDCF1FV3X29B7ovSTXEJaySImdz8CCY5C4","AIzaSyAfKn8uTZOmWcDoVTY_N9YzvZLityhUdWk","AIzaSyDW7sMKMKlQ9y8bar1qWzvqtslJEZAwXq4","AIzaSyAefWW8GFn8wsPMX6LIjusZ5VCa5RWPM3Y","AIzaSyDwVLg6iO8LhuNf3_vROSrM9mBFAXVOVD4","AIzaSyD9mGUGDQNObKfuZUUZTP4ohmxX0tBS1co"]

driver = ""

def setup():
    global driver
    driver = webdriver.Firefox()
    driver.get("http://www.gps-coordinates.net")

   
def sele(address):
    global driver
    query = driver.find_element_by_id('address')
    query.clear()
    query.send_keys(address)
    query.send_keys(Keys.RETURN)
    driver.implicitly_wait(20)
    time.sleep(1)
    query.send_keys(Keys.DOWN) #To solve in case there is autocomplete by Google
    query.send_keys(Keys.RETURN)
    button = driver.find_element_by_xpath("//*[@id='wrap']/div[2]/div[4]/div[1]/form[1]/div[2]/div/button")
    button.click()
    time.sleep(1)
    driver.implicitly_wait(20)
    try:
        alert=driver.switch_to_alert()
        alert.accept()
        latlong = ['Fail','Fail']
        return latlong
    except Exception :
        latraw = driver.find_element_by_id('latitude')
        longraw = driver.find_element_by_id('longitude')
        newquery = driver.find_element_by_id('address')
        lat = latraw.get_attribute('value')
        long = longraw.get_attribute('value')
        query = newquery.get_attribute('value')
        return [lat,long,query]

def wrapperWebParse(inputad,outputad,k):
    InFile = openpyxl.load_workbook(inputad)
    Sheet = InFile['Sheet1']
    workbook = xlsxwriter.Workbook(outputad)
    worksheet = workbook.add_worksheet()
    TotalLength = Sheet.max_row
    ProgressChecker = 0
    worksheet.write(0,0,"Address")
    worksheet.write(0,1,"Latitude")
    worksheet.write(0,2,"Longitude")
    worksheet.write(0,3,"Reverse Address")

    for i in range(1, TotalLength+1):
        ProgressChecker = ProgressChecker + 1
        addtext = Sheet[k+str(i)].value
        try:
            latlong = sele(addtext)
            worksheet.write(i,0,addtext)
            worksheet.write(i,1,latlong[0])
            worksheet.write(i,2,latlong[1])
            worksheet.write(i,3,latlong[2])
            print("Progress: ", ProgressChecker, " out of " ,TotalLength)
        except Exception :
            worksheet.write(i,0,addtext)
            worksheet.write(i,1,"Failure Inside Iteration")
            print("Progress: ", ProgressChecker, " out of " ,TotalLength, "WARNING BRUHNG")
            continue
    complete = "complete"
    return complete
    
def wrapper(inputad,outputad,k,api,geocoder):
    InFile = openpyxl.load_workbook(inputad)
    Sheet = InFile['Sheet1']
    TotalLength = Sheet.max_row
    workbook = xlsxwriter.Workbook(outputad)
    worksheet = workbook.add_worksheet()
    ProgressChecker = 1
    apiind = 0

    if geocoder == "Google":
        g = geocoders.GoogleV3(api_key = api[apiind])
    elif geocoder == "Bing":
        g = geocoders.Bing(api_key = api[apiind])
    elif geocoder == "GeocodeFarm":
        g = geocoders.GeocodeFarm()
    elif geocoder == "GeoNames":
        g = geocoders.GeoNames()
    elif geocoder == "Nominatim":
        g = geocoders.Nominatim()
    elif geocoder == "OpenMapQuest":
        g = geocoders.OpenMapQuest()
    elif geocoder == "Yandex":
        g = geocoders.Yandex()
    else:
        raise ValueError('represents a hidden bug, do not catch this')
    worksheet.write(0,0,"Address")
    worksheet.write(0,1,"Latitude")
    worksheet.write(0,2,"Longitude")
    worksheet.write(0,3,"Reverse Address")
    #Geocoding
    for i in range(1,TotalLength+1):
        if api != "":
            if i % 2500 == 0:
                apiind = apiind + 1
        ProgressChecker = ProgressChecker + 1
        try:
            inputAddress = Sheet[k+str(i)].value
            location = g.geocode(inputAddress, timeout=40)
            worksheet.write(i,0,inputAddress)
            worksheet.write(i,1,location.latitude)
            worksheet.write(i,2,location.longitude)
            worksheet.write(i,3,location.address)
            print("Progress: "+str( ProgressChecker)+ " out of " +str(TotalLength))
        except Exception :
            worksheet.write(i,0,inputAddress)
            worksheet.write(i,1,"Fail")
            print("Progress: "+str( ProgressChecker)+ " out of " +str(TotalLength)+"FAIL")
            continue

def lengthidentify(inputad):
    InFile = openpyxl.load_workbook(inputad)
    Sheet = InFile['Sheet1']
    TotalLength = Sheet.max_row,
    return TotalLength[0]
    
    
def main():
    msg= "Welcome to Geocoder v1.0!"
    title = "Geocoder v1.0"
    if easygui.ccbox(msg, title):
        pass
    else:
        sys.exit(0)

    msg2 = "To start, we will need you to specify the EXCEL FILE where all the addresses is! Make sure that you only choose XLSX file in the next step. Make sure that inside this excel file, all of the addresses is put under the same column, and make sure all the data is put in SHEET 1"
    if easygui.ccbox(msg2, title):
        pass
    else:
        sys.exit(0)


    inputad = easygui.fileopenbox(title)

    while inputad.endswith("xlsx") != True:
        msg3 = "The file that you choose is not XLSX file type. Please Choose Again."
        easygui.msgbox(msg3,title)
        inputad = easygui.fileopenbox(title)

    inputad = os.path.normpath(inputad)


    msg4 = "Which column does the adresses exist?"
    choices = ["A","B","C","D","E","others"]
    finalchoice = easygui.choicebox(msg4,title,choices)

    if finalchoice == "others":
        msg5 = "Write Down the column name where the adresses exists."
        finalchoice = easygui.enterbox(msg5,title)

    while finalchoice.isalpha() !=True:
        msg5 = "The Column name appears to be incorrect. Write Down the column name where the adresses exists."
        finalchoice = easygui.enterbox(msg5,title)

    k = finalchoice

    msg= "We will analyze your excel file. It might take a few minutes."
    if easygui.ccbox(msg, title):
        pass
    else:
        sys.exit(0)

    msg9 = "Choose the geocoder service that you would like to use"
    choices2 = ["Google","Bing","GeocodeFarm","GeoNames","Nominatim","OpenMapQuest","Yandex","Web Parsing"]
    finalchoice2 = easygui.choicebox(msg9,title,choices2)


    geocoder = finalchoice2
    length = lengthidentify(inputad)

    if geocoder == "Bing":
        needapi = "True"
    elif geocoder == "Google":
        needapi ="True"
    else:
        needapi = "False"

    if needapi == "True":
        apineeded = length // 2500 + 1
        msg6 = "To geocode this, you need " + str(apineeded) +" API Keys. Documentation on how to get API Key can be found at http://geopy.readthedocs.io/en/1.11.0/ or google search"

        if easygui.ccbox(msg6, title):
            pass
        else:
            sys.exit(0)

        apifields = [None] * apineeded
        for i in range(apineeded):
            apifields[i] = "API Key" + str(i+1)

        fieldvalues = easygui.multenterbox(msg,title, apifields)

        while 1:
            if fieldvalues is None: break
            errmsg = ""
            for i in range(len(apifields)):
                if fieldvalues[i].strip() == "":
                    errmsg = "Please fill in missing information"
                    break
            if errmsg == "":
                break # no problems found
            fieldvalues = easygui.multenterbox(errmsg, title, apifields, fieldvalues,geocoder)

        api = fieldvalues

    else:
        api = ""

    outputad = inputad[:-5]+"Geocoder"+geocoder+".xlsx"

    msg= "We will start your geocoding now. It might take a few minutes or hours. We will create a file called " + outputad + ". Make sure that there is no file with the same name, or else we will overwrite it."
    if easygui.ccbox(msg, title):
        pass
    else:
        sys.exit(0)


    msg= "Grab a coffee!:) We will let you know once the geocoding process is finished."
    if easygui.ccbox(msg, title):
        pass
    else:
        sys.exit(0)

    try:
        if geocoder != "Web Parsing":
            wrapper(inputad,outputad,k,api,geocoder)
        else:
            setup()
            wrapperWebParse(inputad,outputad, k)
    except:
        easygui.exceptionbox()

    msg= "Done!"
    if easygui.ccbox(msg, title):
        pass
    else:
        sys.exit(0)

if __name__ = "__main__":
    main()
